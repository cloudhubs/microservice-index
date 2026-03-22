#!/usr/bin/env python3
"""
Generate evaluation metrics for the Scenario Generator.

Produces an Excel file with per-endpoint metrics matching the experiment planning
template layout exactly (Scenario Generator Validity sheet).

Data flow:
  input/endpoints.json + input/components.json
    -> output/auth_vectors/all_vectors.json  (ground truth for authorization paths)
    -> output/llm-scenarios/scenarios_llm_ready.json  (scenario generator output to evaluate)

Usage:
  python generate_metrics.py [--scenarios PATH] [--endpoints PATH] [--vectors PATH] [--out PATH]
"""
from __future__ import annotations

import argparse
import json
import re
from typing import Dict, List, Tuple, Set, Any, Optional


# ---------------------------------------------------------------------------
# IO
# ---------------------------------------------------------------------------

def load_json(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

def _normalize_method(m: str) -> str:
    return (m or "").strip().upper()


def _normalize_path(p: str) -> str:
    p = (p or "").strip()
    if not p.startswith("/"):
        p = "/" + p
    p = re.sub(r"/+", "/", p)
    if len(p) > 1 and p.endswith("/"):
        p = p[:-1]
    return p


def _canonicalize_path(path: str) -> str:
    """Replace any {param} with {} for canonical comparison."""
    return re.sub(r"\{[^}]*\}", "{}", _normalize_path(path))


def _strip_api_prefix(uri: str) -> str:
    """Remove /api/v1/ prefix for display labels."""
    return re.sub(r"^/api/v\d+/", "", uri)


def _replace_path_params_display(uri: str) -> str:
    """Replace {paramName} with {?} for display."""
    return re.sub(r"\{[^}]+\}", "{?}", uri)


def make_endpoint_label(method: str, uri: str) -> str:
    """Create display label like 'GET contactservice/contacts'."""
    short = _strip_api_prefix(uri)
    short = _replace_path_params_display(short)
    return f"{method} {short}"


# ---------------------------------------------------------------------------
# Endpoint indexing
# ---------------------------------------------------------------------------

def build_method_path_index(endpoints_json: dict) -> Tuple[Dict[str, List[str]], Dict[str, dict]]:
    """Returns (method+uri -> [endpoint_ids], id -> endpoint_obj)."""
    id_to_endpoint: Dict[str, dict] = endpoints_json.get("endpoints", {})
    key_to_ids: Dict[str, List[str]] = {}
    for eid, ep in id_to_endpoint.items():
        method = _normalize_method(ep.get("httpMethod") or "")
        uri = _normalize_path(ep.get("fullUri") or ep.get("simplifiedUri") or "")
        if method and uri:
            k = f"{method}|{uri}"
            key_to_ids.setdefault(k, []).append(eid)
    return key_to_ids, id_to_endpoint


def resolve_endpoint_id(endpoints_json: dict, method: str, uri: str) -> Optional[str]:
    key_to_ids, _ = build_method_path_index(endpoints_json)
    k = f"{_normalize_method(method)}|{_normalize_path(uri)}"
    ids = key_to_ids.get(k, [])
    return ids[0] if len(ids) == 1 else None


def get_roles_universe(vectors_json: dict) -> List[str]:
    md = vectors_json.get("metadata", {})
    roles = md.get("roles", [])
    if roles:
        return list(roles)
    for _, v in vectors_json.get("vectors", {}).items():
        r = v.get("permission_matrix", {}).get("roles", [])
        if r:
            return list(r)
    return []


# ---------------------------------------------------------------------------
# Target matching
# ---------------------------------------------------------------------------

def find_matching_scenarios(
    scenarios: List[dict], method: str, uri_pattern: str
) -> List[dict]:
    """Find all scenarios matching a target by endpoint field (method + URI pattern)."""
    canon = _canonicalize_path(uri_pattern)
    matched = []
    for s in scenarios:
        if _normalize_method(s.get("method", "")) != _normalize_method(method):
            continue
        s_canon = _canonicalize_path(s.get("endpoint", ""))
        if s_canon == canon:
            matched.append(s)
    return matched


def find_all_related_scenarios(
    scenarios: List[dict], method: str, uri_pattern: str,
    target_eids: List[str]
) -> List[dict]:
    """Find ALL scenarios related to a target endpoint.

    Includes:
    1. Direct match: scenario endpoint == target (same method + URI)
    2. Chain match: target endpoint ID appears in scenario's chain_permissions
       (the target is a downstream endpoint in another vector's chain)
    """
    canon = _canonicalize_path(uri_pattern)
    norm_method = _normalize_method(method)
    eid_set = set(target_eids)
    seen_ids: Set[str] = set()
    matched = []

    for s in scenarios:
        sid = s.get("scenario_id", "")
        if sid in seen_ids:
            continue

        # Direct match by endpoint field
        if (_normalize_method(s.get("method", "")) == norm_method
                and _canonicalize_path(s.get("endpoint", "")) == canon):
            seen_ids.add(sid)
            matched.append(s)
            continue

        # Chain match: target endpoint ID appears in chain_permissions
        if eid_set:
            cp = s.get("chain_permissions", {}) or {}
            if eid_set & set(cp.keys()):
                seen_ids.add(sid)
                matched.append(s)

    return matched


def find_endpoint_ids_for_pattern(
    endpoints_json: dict, method: str, uri_pattern: str
) -> List[str]:
    """Find all endpoint IDs matching a target pattern."""
    canon = _canonicalize_path(uri_pattern)
    _, id_to_ep = build_method_path_index(endpoints_json)
    out = []
    for eid, ep in id_to_ep.items():
        ep_method = _normalize_method(ep.get("httpMethod") or "")
        ep_uri = ep.get("simplifiedUri") or ep.get("fullUri") or ""
        if ep_method == _normalize_method(method) and _canonicalize_path(ep_uri) == canon:
            out.append(eid)
    return out


# ---------------------------------------------------------------------------
# Ground truth derivations
# ---------------------------------------------------------------------------

def derive_entrypoint_allowed_roles(endpoint_obj: dict, roles_universe: List[str]) -> Set[str]:
    auth = endpoint_obj.get("authorization", {}) or {}
    authentication = endpoint_obj.get("authentication", {}) or {}
    if bool(auth.get("public", False)):
        return set(roles_universe)
    required = auth.get("requiredRoles", []) or []
    if not required and bool(auth.get("requiresAuthentication", False)):
        return set(roles_universe)
    if not required and bool(authentication.get("required", False)) and not bool(authentication.get("permitAll", False)):
        return set(roles_universe)
    return set(r for r in required if r in roles_universe)


def derive_expected_status(endpoint_obj: dict, roles_universe: List[str]) -> Dict[str, str]:
    allowed = derive_entrypoint_allowed_roles(endpoint_obj, roles_universe)
    return {r: "2xx" if r in allowed else "403" for r in roles_universe}


def _normalize_status(s: Any) -> str:
    s = str(s).strip()
    if s in ("200", "201", "204"):
        return "2xx"
    if s.startswith("2") and s.endswith("xx"):
        return "2xx"
    return s


def derive_inconsistencies_from_vector(vector: dict) -> Set[Tuple[str, str, str, str]]:
    """
    Derive ground-truth inconsistencies from a single vector's permission matrix.
    Matrix orientation: matrix[endpoint_idx][role_idx].
    Returns set of (type, role, root_endpoint_id, downstream_endpoint_id).
    """
    gt: Set[Tuple[str, str, str, str]] = set()
    pm = vector.get("permission_matrix", {}) or {}
    roles = pm.get("roles", []) or []
    endpoints = pm.get("endpoints", []) or []
    matrix = pm.get("matrix", []) or []

    if not roles or len(endpoints) < 2 or len(matrix) != len(endpoints):
        return gt

    root_perms = matrix[0]
    if len(root_perms) != len(roles):
        return gt

    for j in range(1, len(endpoints)):
        if j >= len(matrix):
            continue
        down_perms = matrix[j]
        if len(down_perms) != len(roles):
            continue
        for r_i, role in enumerate(roles):
            root_val = root_perms[r_i]
            down_val = down_perms[r_i]
            if root_val == 0 and down_val == 1:
                gt.add(("OverPermissiveDownstream", role, endpoints[0], endpoints[j]))
            elif root_val == 1 and down_val == 0:
                gt.add(("UnderPermissiveDownstream", role, endpoints[0], endpoints[j]))

    return gt


def derive_all_inconsistencies_from_vectors(vectors_json: dict) -> Set[Tuple[str, str, str, str]]:
    gt: Set[Tuple[str, str, str, str]] = set()
    for _, v in vectors_json.get("vectors", {}).items():
        if len(v.get("call_chain", []) or []) > 1:
            gt |= derive_inconsistencies_from_vector(v)
    return gt


def has_policy_exposure_ground_truth(scenario: dict, endpoint_obj: dict) -> bool:
    """
    PolicyExposure = endpoint is public AND handles sensitive data (PII/Financial).
    Uses sensitivity_type from source code analysis (independent of LLM).
    """
    auth = endpoint_obj.get("authorization", {}) or {}
    if not bool(auth.get("public", False)):
        return False
    sensitivity = scenario.get("sensitivity_type", "")
    if sensitivity in ("FINANCIAL", "PII"):
        return True
    if scenario.get("handles_pii", False):
        return True
    return False


# ---------------------------------------------------------------------------
# Per-target metric computation
# ---------------------------------------------------------------------------

def compute_target_metrics(
    target_method: str,
    target_uri: str,
    scenarios: List[dict],
    endpoints_json: dict,
    vectors_json: dict,
    all_roles: List[str],
) -> dict:
    """Compute all RQ1+RQ2 metrics for one target endpoint."""

    matched = find_matching_scenarios(scenarios, target_method, target_uri)
    label = make_endpoint_label(target_method, target_uri)
    roles_set = set(all_roles)
    all_vectors = vectors_json.get("vectors", {})
    _, id_to_ep = build_method_path_index(endpoints_json)

    # Resolve endpoint IDs for this target
    eids = find_endpoint_ids_for_pattern(endpoints_json, target_method, target_uri)
    eid = eids[0] if len(eids) == 1 else None

    # Try scenario-based fallback
    if not eid and matched:
        sid = matched[0].get("scenario_id", "")
        candidate = sid[4:] if sid.startswith("scn_") else sid
        if candidate in all_vectors:
            eid = candidate

    endpoint_obj = id_to_ep.get(eid) if eid else None
    vector = all_vectors.get(eid) if eid else None

    # Ground truth inconsistencies for this endpoint (from vectors)
    gt_incs: Set[Tuple[str, str, str, str]] = set()
    if vector and len(vector.get("call_chain", []) or []) > 1:
        gt_incs = derive_inconsistencies_from_vector(vector)

    # PolicyExposure ground truth
    has_pe = False
    if endpoint_obj and matched:
        has_pe = has_policy_exposure_ground_truth(matched[0], endpoint_obj)

    # ===== RQ1: Completeness =====

    # Role Coverage
    roles_covered = set()
    for s in matched:
        for k in ("allowed_roles", "denied_roles", "unknown_roles"):
            for r in (s.get(k, []) or []):
                if r in roles_set:
                    roles_covered.add(r)
        for r in (s.get("expected_status_by_role", {}) or {}).keys():
            if r in roles_set:
                roles_covered.add(r)

    re_combinations = len(roles_covered)
    total_combinations = len(all_roles)

    # Path Coverage
    # IMPORTANT: in this dataset each vector represents a single call-chain (a single path),
    # even if it contains many calls/nodes. Previously we used len(call_chain), which is
    # the number of nodes/calls.
    total_paths = 1

    entry_scns = [s for s in matched if s.get("type") == "EntryPointAuthorization"]
    down_scns = [s for s in matched if s.get("type") == "DownstreamPolicyConsistency"]

    if down_scns:
        paths_defined = total_paths
    elif entry_scns:
        paths_defined = 1
    else:
        paths_defined = 0

    # Inconsistency Detection Rate (only downstream scenarios can define policy inconsistencies)
    scenario_inc_count = sum(len(s.get("policy_inconsistencies", []) or []) for s in down_scns)
    gt_inc_count = len(gt_incs) + (1 if has_pe else 0)

    # Entry/Downstream counts
    entry_cases = len(entry_scns)
    down_cases = len(down_scns)

    # Positive/Negative
    positive_cases = sum(len(s.get("allowed_roles", []) or []) for s in matched)
    negative_cases = sum(len(s.get("denied_roles", []) or []) for s in matched)

    # ===== RQ2: Correctness =====

    # Correct Expected Outcome (per scenario)
    correct_outcome = 0
    total_outcome = 0
    if endpoint_obj:
        gt_status = derive_expected_status(endpoint_obj, all_roles)
        for s in matched:
            observed = s.get("expected_status_by_role", {}) or {}
            for r in all_roles:
                total_outcome += 1
                if _normalize_status(observed.get(r, "")) == _normalize_status(gt_status.get(r, "")):
                    correct_outcome += 1

    # Correct Role Assignment (per role across each scenario)
    correct_role = 0
    total_role = 0
    if endpoint_obj:
        gt_allowed = derive_entrypoint_allowed_roles(endpoint_obj, all_roles)
        gt_denied = roles_set - gt_allowed
        for s in matched:
            if s.get("type") == "EntryPointAuthorization":
                s_allowed = set(s.get("allowed_roles", []) or []) & roles_set
                s_denied = set(s.get("denied_roles", []) or []) & roles_set
                for r in all_roles:
                    total_role += 1
                    if ((r in s_allowed) == (r in gt_allowed)
                            and (r in s_denied) == (r in gt_denied)):
                        correct_role += 1
            elif s.get("type") == "DownstreamPolicyConsistency" and vector:
                chain_perm = s.get("chain_permissions", {}) or {}
                pm = vector.get("permission_matrix", {}) or {}
                pm_roles = pm.get("roles", []) or []
                pm_endpoints = pm.get("endpoints", []) or []
                pm_matrix = pm.get("matrix", []) or []

                if pm_roles and pm_endpoints and pm_matrix:
                    role_idx = {r: i for i, r in enumerate(pm_roles)}
                    ep_idx = {e: j for j, e in enumerate(pm_endpoints)}
                    per_role_ok = {r: True for r in all_roles}
                    for node_id, node_perm in chain_perm.items():
                        if node_id not in ep_idx:
                            continue
                        j = ep_idx[node_id]
                        if j >= len(pm_matrix):
                            continue
                        row = pm_matrix[j]  # matrix[endpoint_idx][role_idx]
                        s_a = set(node_perm.get("allowed_roles", []) or []) & set(pm_roles)
                        s_d = set(node_perm.get("denied_roles", []) or []) & set(pm_roles)
                        for r in all_roles:
                            ri = role_idx.get(r)
                            if ri is None or ri >= len(row):
                                continue
                            gt_is_allowed = row[ri] == 1
                            gt_is_denied = not gt_is_allowed
                            if ((r in s_a) != gt_is_allowed) or ((r in s_d) != gt_is_denied):
                                per_role_ok[r] = False
                    for r in all_roles:
                        total_role += 1
                        if per_role_ok.get(r, False):
                            correct_role += 1

    # Correct Parameter Specification (per scenario)
    correct_param = 0
    total_param = 0
    if endpoint_obj:
        for s in matched:
            ep_body = endpoint_obj.get("requestBodyDetail", {}) or {}
            ep_body_type = ep_body.get("type", "")
            ep_path_params = endpoint_obj.get("pathParameterDetails", []) or []

            ok = True
            for pp in ep_path_params:
                if pp.get("name"):
                    s_pp = s.get("pathParameterDetails", []) or []
                    if pp["name"] not in {p.get("name", "") for p in s_pp}:
                        ok = False
                        break
            if ok and ep_body_type:
                s_body = s.get("data", {}).get("request_body_type", "")
                s_entity = (s.get("entity_schema", {}) or {}).get("entity_name", "")
                if s_body != ep_body_type and s_entity != ep_body_type:
                    ok = False
            total_param += len(all_roles)
            if ok:
                correct_param += len(all_roles)

    # Confusion Matrix (template convention) computed PER DOWNSTREAM SCENARIO.
    # We classify a scenario as "GT inconsistent" if any (root, downstream) pair in its
    # chain intersects a ground-truth inconsistent pair derived from the vector.
    # This avoids:
    # - mixing EntryPointAuthorization scenarios into inconsistency metrics
    # - using a single gt_has_inconsistency boolean for the whole endpoint
    gt_pairs = {(u, d) for (_, _, u, d) in gt_incs}

    tn_val = 0  # correctly captures inconsistency
    fp_val = 0  # missed inconsistency
    fn_val = 0  # incorrectly captures inconsistencies
    tp_val = 0  # tests consistency

    for s in down_scns:
        chain_perm = s.get("chain_permissions", {}) or {}
        chain_ids = [k for k in chain_perm.keys() if isinstance(k, str) and k]

        predicted_pairs: Set[Tuple[str, str]] = set()
        for inc in (s.get("policy_inconsistencies", []) or []):
            if not isinstance(inc, dict):
                continue
            details = inc.get("details", {}) or {}
            up = details.get("upstream_endpoint") or details.get("upstream") or ""
            down = details.get("downstream_endpoint") or details.get("downstream") or ""
            if up and down:
                predicted_pairs.add((up, down))

        # Build scenario chain pairs as (root, downstream)
        root_id = next(iter(predicted_pairs))[0] if predicted_pairs else (chain_ids[0] if chain_ids else "")
        scenario_pairs: Set[Tuple[str, str]] = set(predicted_pairs)
        if root_id:
            for d_id in chain_ids[1:]:
                if d_id != root_id:
                    scenario_pairs.add((root_id, d_id))

        predicted_inconsistent = bool(predicted_pairs)
        actual_inconsistent = bool(gt_pairs) and any(p in gt_pairs for p in scenario_pairs)

        if actual_inconsistent and predicted_inconsistent:
            tn_val += 1
        elif actual_inconsistent and not predicted_inconsistent:
            fp_val += 1
        elif not actual_inconsistent and predicted_inconsistent:
            fn_val += 1
        else:
            tp_val += 1

    return {
        "label": label,
        "matched_scenarios": len(matched),
        # RQ1
        "re_combinations": re_combinations,
        "total_combinations": total_combinations,
        "paths_defined": paths_defined,
        "total_paths": total_paths,
        "inc_defined": scenario_inc_count,
        "inc_in_path": gt_inc_count,
        "entry_cases": entry_cases,
        "down_cases": down_cases,
        "positive": positive_cases,
        "negative": negative_cases,
        # RQ2
        "correct_outcome": correct_outcome,
        "total_outcome": total_outcome,
        "correct_role": correct_role,
        "total_role": total_role,
        "correct_param": correct_param,
        "total_param": total_param,
        "tn": tn_val,  # K
        "fp": fp_val,  # L
        "fn": fn_val,  # N
        "tp": tp_val,  # O
    }


def compute_rq3_metrics(scenarios: List[dict]) -> dict:
    """Compute RQ3 Executability metrics (global)."""
    total = len(scenarios)

    valid_count = sum(
        1 for s in scenarios
        if all(k in s for k in ("scenario_id", "type", "endpoint", "method",
                                 "allowed_roles", "expected_status_by_role"))
    )

    complete_count = 0
    for s in scenarios:
        has_all = True
        body_type = s.get("data", {}).get("request_body_type", "")
        entity = s.get("entity_schema", {}) or {}
        if body_type and body_type not in ("", "void"):
            if not (entity.get("fields") or s.get("curlExample")):
                has_all = False
        if has_all:
            complete_count += 1

    seen = set()
    dup_count = 0
    for s in scenarios:
        sig = (s.get("method", ""), s.get("endpoint", ""), s.get("type", ""),
               tuple(sorted(s.get("allowed_roles", []) or [])),
               tuple(sorted(s.get("denied_roles", []) or [])))
        if sig in seen:
            dup_count += 1
        seen.add(sig)

    return {
        "total": total,
        "valid_count": valid_count,
        "complete_count": complete_count,
        "dup_count": dup_count,
    }


# ---------------------------------------------------------------------------
# Excel output (exact template layout)
# ---------------------------------------------------------------------------

def write_excel(out_path: str, target_rows: List[dict], rq3: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Scenario Generator Validity"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    n = len(target_rows)

    # ===================== RQ1: Completeness Metrics =====================

    # Row 2: title
    ws.cell(row=2, column=1, value="Round 1 (System AS IT IS)").font = bold

    # Row 3: section label
    ws.cell(row=3, column=2, value="Completeness Metrics").font = bold

    # Row 4: group headers
    groups_rq1 = [
        (2, 4, "Role Coverage"),
        (5, 7, "Path Coverage"),
        (8, 10, "Inconsistencies Detection Rate "),
        (11, 13, "Entry Point Testing Ratio"),
        (14, 16, "Postive to Negative Ratio"),
    ]
    for sc, ec, title in groups_rq1:
        ws.cell(row=4, column=sc, value=title).font = bold
        ws.cell(row=4, column=sc).alignment = center
        if ec > sc:
            ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)

    # Row 5: column headers
    headers_rq1 = {
        1: "Target Endpoint", 2: "R-E Combinations", 3: "Total Combinations",
        4: "Role Coverage %", 5: "Paths defined", 6: "Total Paths",
        7: "Path Coverage %", 8: "Inconsistencies Defined",
        9: "All Inconsistencies in Path", 10: "Detection Rate",
        11: "Entry Point Cases", 12: "Down Stream Cases", 13: "Testing Ratio",
        14: "Positive Cases", 15: "Negative Cases", 16: "Ratio",
    }
    for c, h in headers_rq1.items():
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = bold
        cell.alignment = center

    # Rows 6..6+n-1: data
    for i, row_data in enumerate(target_rows):
        r = 6 + i
        ws.cell(row=r, column=1, value=row_data["label"])
        ws.cell(row=r, column=2, value=row_data["re_combinations"])
        ws.cell(row=r, column=3, value=row_data["total_combinations"])
        ws.cell(row=r, column=4, value=f"=(B{r}/C{r})*100")
        ws.cell(row=r, column=5, value=row_data["paths_defined"])
        ws.cell(row=r, column=6, value=row_data["total_paths"])
        ws.cell(row=r, column=7, value=f"=IF(F{r}>0,(E{r}/F{r})*100,0)")
        ws.cell(row=r, column=8, value=row_data["inc_defined"])
        ws.cell(row=r, column=9, value=row_data["inc_in_path"])
        ws.cell(row=r, column=10, value=f"=IF(I{r}>0,(H{r}/I{r})*100,0)")
        ws.cell(row=r, column=11, value=row_data["entry_cases"])
        ws.cell(row=r, column=12, value=row_data["down_cases"])
        ws.cell(row=r, column=13, value=f"=IF(L{r}+K{r}>0,K{r}/(K{r}+L{r}),0)")
        ws.cell(row=r, column=14, value=row_data["positive"])
        ws.cell(row=r, column=15, value=row_data["negative"])
        ws.cell(row=r, column=16, value=f"=IF(O{r}>0,O{r}/N{r},0)")

    # ===================== RQ2: Correctness Metrics =====================

    # Row after data + 3 gap rows
    rq2_section_row = 6 + n + 3  # e.g., row 22 for 13 rows

    ws.cell(row=rq2_section_row, column=2, value="Correctness Metrics").font = bold

    # Group headers
    rq2_group_row = rq2_section_row + 1
    groups_rq2 = [
        (2, 4, "Correct Expected Outcome Rate"),
        (5, 7, "Correct Role Assignment Rate"),
        (8, 10, "Correct Parameter Specification Rate "),
        (11, 13, "True Negative Rate"),
        (14, 16, "False Negative Rate"),
        (17, 19, "False Positive Rate"),
        (20, 20, "Precision"),
        (21, 21, "Recall"),
        (22, 22, "F1 Score"),
    ]
    for sc, ec, title in groups_rq2:
        ws.cell(row=rq2_group_row, column=sc, value=title).font = bold
        ws.cell(row=rq2_group_row, column=sc).alignment = center
        if ec > sc:
            ws.merge_cells(start_row=rq2_group_row, start_column=sc,
                           end_row=rq2_group_row, end_column=ec)

    # Column headers
    rq2_header_row = rq2_group_row + 1
    headers_rq2 = {
        1: "Target Endpoint",
        2: "Cases with Correct Expected Status Codes ",
        3: "Total Cases",
        4: "Correct Outcome %",
        5: "Cases with correct role-permission mapping",
        6: "Total Cases",
        7: "Correct Role Assignment %",
        8: "Cases with Valid Parameters",
        9: "Total Cases",
        10: "Parameter Specification Rate",
        11: "Cases that correctly captures inconsistency (TN)",
        12: "Cases that missed inconsistency (FP)",
        13: "TNR",
        14: "Cases that incorrectly captures inconsistencies (FN)",
        15: "Cases that tests consistency (TP) ",
        16: "FNR",
        17: "Cases that missed inconsistency (FP)",
        18: "Cases that correctly captures inconsistency (TN)",
        19: "FPR",
        20: "Precision",
        21: "Recall",
        22: "F1 Score",
    }
    for c, h in headers_rq2.items():
        cell = ws.cell(row=rq2_header_row, column=c, value=h)
        cell.font = bold
        cell.alignment = center

    # Data rows
    rq2_data_start = rq2_header_row + 1
    for i, row_data in enumerate(target_rows):
        r = rq2_data_start + i
        ws.cell(row=r, column=1, value=row_data["label"])
        ws.cell(row=r, column=2, value=row_data["correct_outcome"])
        ws.cell(row=r, column=3, value=row_data["total_outcome"])
        ws.cell(row=r, column=4, value=f"=IF(C{r}>0,(B{r}/C{r})*100,0)")
        ws.cell(row=r, column=5, value=row_data["correct_role"])
        ws.cell(row=r, column=6, value=row_data["total_role"])
        ws.cell(row=r, column=7, value=f"=(E{r}/F{r})*100")
        ws.cell(row=r, column=8, value=row_data["correct_param"])
        ws.cell(row=r, column=9, value=row_data["total_param"])
        ws.cell(row=r, column=10, value=f"=IF(I{r}>0,(H{r}/I{r})*100,0)")
        ws.cell(row=r, column=11, value=row_data["tn"])
        ws.cell(row=r, column=12, value=row_data["fp"])
        ws.cell(row=r, column=13, value=f"=IF(K{r}+L{r}>0,K{r}/(K{r}+L{r}),0)")
        ws.cell(row=r, column=14, value=row_data["fn"])
        ws.cell(row=r, column=15, value=row_data["tp"])
        ws.cell(row=r, column=16, value=f"=IF(N{r}+O{r}>0,N{r}/(N{r}+O{r}),0)")
        ws.cell(row=r, column=17, value=row_data["fp"])
        ws.cell(row=r, column=18, value=row_data["tn"])
        ws.cell(row=r, column=19, value=f"=IF(Q{r}+R{r}>0,Q{r}/(Q{r}+R{r}),0)")
        ws.cell(row=r, column=20, value=f"=IF(K{r}+N{r}>0,K{r}/(K{r}+N{r}),0)")
        ws.cell(row=r, column=21, value=f"=IF(K{r}+Q{r}>0,K{r}/(K{r}+Q{r}),0)")
        ws.cell(row=r, column=22, value=f"=IF(T{r}+U{r}>0,2*(T{r}*U{r})/(T{r}+U{r}),0)")

    # ===================== RQ3: Executability Metrics =====================

    rq3_section_row = rq2_data_start + n + 3

    ws.cell(row=rq3_section_row, column=2, value="Executability Metrics").font = bold

    rq3_group_row = rq3_section_row + 1
    groups_rq3 = [
        (2, 4, "Valid Output Rate"),
        (5, 7, "Parameter Completeness"),
        (8, 10, "Duplicate Rate"),
    ]
    for sc, ec, title in groups_rq3:
        ws.cell(row=rq3_group_row, column=sc, value=title).font = bold
        ws.cell(row=rq3_group_row, column=sc).alignment = center
        if ec > sc:
            ws.merge_cells(start_row=rq3_group_row, start_column=sc,
                           end_row=rq3_group_row, end_column=ec)

    rq3_header_row = rq3_group_row + 1
    headers_rq3 = {
        1: "",
        2: "Valid JSON Outputs", 3: "All Outputs", 4: "Valid Output Rate %",
        5: "Scenarios with all required parameters", 6: "Total Scenarios",
        7: "Parameter Completeness %",
        8: "Duplicate Scenarios", 9: "Total Scenarios", 10: "Duplicate Rate %",
    }
    for c, h in headers_rq3.items():
        cell = ws.cell(row=rq3_header_row, column=c, value=h)
        cell.font = bold
        cell.alignment = center

    rq3_data_row = rq3_header_row + 1
    ws.cell(row=rq3_data_row, column=1, value="All Scenarios")
    ws.cell(row=rq3_data_row, column=2, value=rq3["valid_count"])
    ws.cell(row=rq3_data_row, column=3, value=rq3["total"])
    ws.cell(row=rq3_data_row, column=4, value=f"=IF(C{rq3_data_row}>0,(B{rq3_data_row}/C{rq3_data_row})*100,0)")
    ws.cell(row=rq3_data_row, column=5, value=rq3["complete_count"])
    ws.cell(row=rq3_data_row, column=6, value=rq3["total"])
    ws.cell(row=rq3_data_row, column=7, value=f"=IF(F{rq3_data_row}>0,(E{rq3_data_row}/F{rq3_data_row})*100,0)")
    ws.cell(row=rq3_data_row, column=8, value=rq3["dup_count"])
    ws.cell(row=rq3_data_row, column=9, value=rq3["total"])
    ws.cell(row=rq3_data_row, column=10, value=f"=IF(I{rq3_data_row}>0,(H{rq3_data_row}/I{rq3_data_row})*100,0)")

    # ===================== RQ4: placeholder =====================

    rq4_row = rq3_data_row + 3
    ws.cell(row=rq4_row, column=2, value="Detection Effectiveness (RQ4 - requires test execution)").font = bold

    rq4_header_row = rq4_row + 1
    for c, h in {1: "", 2: "Detected Inconsistencies", 3: "Known Inconsistencies",
                  4: "Inconsistency Detection Rate %"}.items():
        ws.cell(row=rq4_header_row, column=c, value=h).font = bold
        ws.cell(row=rq4_header_row, column=c).alignment = center

    rq4_data_row = rq4_header_row + 1
    ws.cell(row=rq4_data_row, column=1, value="After Test Execution")
    ws.cell(row=rq4_data_row, column=2, value="(pending)")
    ws.cell(row=rq4_data_row, column=3, value="(pending)")
    ws.cell(row=rq4_data_row, column=4, value="N/A")

    # ===================== Column widths =====================
    ws.column_dimensions["A"].width = 48
    for c_letter in "BCDEFGHIJKLMNOPQRSTUV":
        ws.column_dimensions[c_letter].width = 16

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

TARGETS: List[Tuple[str, str]] = [
    ("GET", "/backoffice/ratings/latest/{count}"),
    ("GET", "/storefront/customer/profile"),
    ("GET", "/storefront/cart/items"),
    ("POST", "/backoffice/products"),
    ("POST", "/storefront/ratings"),
    ("POST", "/storefront/sampledata"),
    ("DELETE", "/backoffice/ratings/{id}"),
    ("DELETE", "/ApiConstant.WAREHOUSE_URL/{id}"),
    ("PUT", "/backoffice/products/subtract-quantity"),
    ("PUT", "/Constants.ApiConstant.STATE_OR_PROVINCES_URL/{id}"),
]


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate scenario generator evaluation metrics")
    # round_base = "train-ticket-aitest/master"
    # round_base = "train-ticket-aitest/first-inconsistency-injection/"
    # round_base = "train-ticket-aitest/second-inconsistency-injection/"
    # round_base = "yas/master"
    # round_base = "yas/first-injection"
    round_base = "yas/second-injection/"
    ap.add_argument("--scenarios", default=f"{round_base}/scenarios_llm_ready.json")
    ap.add_argument("--endpoints", default=f"{round_base}/endpoints.json")
    ap.add_argument("--vectors", default=f"{round_base}/all_vectors.json")
    ap.add_argument("--out", default=f"{round_base}/output/metrics.xlsx")
    args = ap.parse_args()

    scenarios = load_json(args.scenarios)
    endpoints_json = load_json(args.endpoints)
    vectors_json = load_json(args.vectors)

    all_roles = get_roles_universe(vectors_json)

    print(f"Total scenarios: {len(scenarios)}")
    print(f"Roles: {all_roles}")
    print(f"Targets: {len(TARGETS)}")

    target_rows: List[dict] = []
    warnings: List[str] = []
    for method, uri in TARGETS:
        row = compute_target_metrics(method, uri, scenarios, endpoints_json,
                                      vectors_json, all_roles)
        target_rows.append(row)

        label = row["label"]
        matched_n = row.get("matched_scenarios", 0)
        entry = row["entry_cases"]
        down = row["down_cases"]
        inc_d = row["inc_defined"]
        inc_gt = row["inc_in_path"]
        print(f"  {label}: E={entry} D={down} inc_def={inc_d} inc_gt={inc_gt} "
              f"TN={row['tn']} FP={row['fp']} FN={row['fn']} TP={row['tp']}")

        # ----------------------
        # Sanity checks (no assumptions about minimum scenarios)
        # ----------------------
        if matched_n != (entry + down):
            warnings.append(
                f"[WARN] {label}: matched_scenarios({matched_n}) != entry_cases({entry}) + down_cases({down})"
            )

        if inc_d > 0 and inc_gt == 0:
            warnings.append(
                f"[WARN] {label}: inc_defined({inc_d}) > 0 but inc_in_path({inc_gt}) == 0 (scenario marked inconsistency where GT has none)"
            )

        if down == 0 and any(row[k] > 0 for k in ("tn", "fp", "fn", "tp")):
            warnings.append(
                f"[WARN] {label}: down_cases==0 but confusion values are non-zero (tn/fp/fn/tp)"
            )

        if row.get("total_paths", 0) == 0 and row.get("paths_defined", 0) > 0:
            warnings.append(
                f"[WARN] {label}: paths_defined({row.get('paths_defined')}) > 0 but total_paths==0"
            )

    rq3 = compute_rq3_metrics(scenarios)

    print(f"\n--- Summary ---")
    total_tn = sum(r["tn"] for r in target_rows)
    total_fp = sum(r["fp"] for r in target_rows)
    total_fn = sum(r["fn"] for r in target_rows)
    total_tp = sum(r["tp"] for r in target_rows)
    print(f"Confusion (template convention): TN={total_tn} FP={total_fp} FN={total_fn} TP={total_tp}")
    print(f"RQ3: Valid={rq3['valid_count']}/{rq3['total']} "
          f"Complete={rq3['complete_count']}/{rq3['total']} "
          f"Duplicates={rq3['dup_count']}/{rq3['total']}")

    if warnings:
        print("\n--- Sanity check warnings ---")
        for w in warnings:
            print(w)
        print(f"Total warnings: {len(warnings)}")
    else:
        print("\n--- Sanity check warnings ---\n(none)")

    write_excel(args.out, target_rows, rq3)
    print(f"\nSaved: {args.out}")


if __name__ == "__main__":
    main()
